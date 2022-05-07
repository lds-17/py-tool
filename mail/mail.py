from openpyxl import Workbook
from openpyxl import load_workbook
import sys
import os

# #sheet名称
# sheet.title
# #最大行和最大列
# sheet.max_row
# sheet.max_column
# #行列生成器
# sheet.rows #为行生成器, 里面是每一行的cell对象，由一个tuple包裹。
# sheet.columns #为列生成器, 里面是每一列的cell对象，由一个tuple包裹。

# 可以使用list(sheet.rows)[0].value 类似方法来获取数据，或
# for row in sheet.rows:
#     for cell in row:
#         print(cell.value)
# 来遍历值,或值生成器 sheet.values 仅遍历值

# #按行完整遍历表
# for row in sheet:
#     for cell in row:
#         print(cell.value,end=",")#默认end是换行

# #按行指定区域遍历表
# for row in sheet.iter_rows(min_row=3,max_row=4,min_col=3,max_col=20):
#     for cell in col:
#         print(cell.value, end=",")

# #按列遍历表
# for col in sheet.columns:
#     for cell in col:
#         print(cell.value, end=",")

# #按列指定区域遍历表
# for col in sheet.iter_cols(min_col=3,max_col=4,min_row=3,max_row=20):
#     for cell in col:
#         print(cell.value, end=",")

# sheet.append(row)  # 添加一行数据

# cell:
# sheet[A]: 数组, 返回A列所有cell
# sheet[1]: 数组, 返回第1行所有cell
# sheet[A1]: cell对象, 第1行第A列
# sheet[A:E]: 数组, A-E列的所有cell
# #返回列
# cell.column
# #返回行
# cell.row
# #返回值
# cell.value
# 注意:如果单元格是使用的公式，则值是公式而不是计算后的值
# #返回单元格格式属性
# cell.number_format
# 默认为General格式
# #单元格样式
# cell.font

class Mail(object):
	def __init__(self, address, title, content):
		super(Mail, self).__init__()
		self.address = address
		self.title = title
		self.content = content
		

class AutoEMail(object):
	"""docstring for AutoEMail"""
	def __init__(self):
		super(AutoEMail, self).__init__()
		cfg = open('./mail.cfg', encoding='utf-8')
		self.smtp = None
		self.port = None
		self.ssl_port = None
		self.code = None
		self.mails = []
		for line in cfg:
			t = line.strip().split(':')
			if len(t) < 2 : continue
			k = t[0].strip()
			v = t[1].strip()
			if k == 'smtp': self.smtp = v
			if k == 'port': self.port = v 
			if k == 'ssl_port': self.ssl_port = v 
			if k == 'code': self.code = v
		if self.smtp == None or self.port == None or self.ssl_port == None or self.code == None : 
			print("mail.cfg error")
			raise("mail.cfg error")

	def pf(self):
		print('smtp: ',self.smtp)
		print('port: ', self.port)
		print('ssl_port: ', self.ssl_port)
		print('code: ', self.code)
		

	def readDir(self):
		files = os.listdir('./')
		files = [x for x in files if (x.lower().endswith('.xlsx') or x.lower().endswith('.xls')) and (not x.startswith('~$'))]
		for file in files:
			self.read(file)

	def read(self, file):
		wb = load_workbook(filename=file, read_only=True)
		print('encoding: ', wb.encoding)
		for sheetName in wb.sheetnames:
			self.doSheet(wb[sheetName])

	def doSheet(self, sheet):
		for row in sheet:
			for cell in row:
				print(cell.value)	

# var = input("Y/N:\n")
# print(var)

Mail = AutoEMail()
Mail.pf()
Mail.readDir()